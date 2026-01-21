"""Utility to load data from Excel files"""
import os
from pathlib import Path
from typing import List

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    import odf.opendocument
    import odf.table
except ImportError:
    odf = None


def load_customers_from_ods() -> List[str]:
    """
    Load customers from the Classeur1.ods file in the data folder
    Returns a list of unique customer names
    """
    data_dir = Path(__file__).resolve().parent.parent.parent.parent / "data"
    ods_file = data_dir / "Classeur1.ods"
    
    if not ods_file.exists():
        print(f"Warning: Customer file not found at {ods_file}")
        return []
    
    try:
        if odf is not None:
            return _load_from_ods(str(ods_file))
        else:
            print("Warning: openpyxl or odfpy not installed, returning empty customer list")
            return []
    except Exception as e:
        print(f"Error loading customers from ODS: {e}")
        return []


def _load_from_ods(file_path: str) -> List[str]:
    """Load customers from ODS file"""
    try:
        doc = odf.opendocument.load(file_path)
        tables = doc.spreadsheet.getElementsByType(odf.table.Table)
        
        customers = set()
        for table in tables:
            rows = table.getElementsByType(odf.table.TableRow)
            for row_idx, row in enumerate(rows):
                if row_idx == 0:  # Skip header
                    continue
                cells = row.getElementsByType(odf.table.TableCell)
                if cells:
                    cell_text = str(cells[0])
                    if cell_text.strip():
                        customers.add(cell_text.strip())
        
        return sorted(list(customers))
    except Exception as e:
        print(f"Error parsing ODS file: {e}")
        return []


def load_customers_from_xlsx() -> List[str]:
    """
    Load customers from an Excel file (for backward compatibility)
    """
    data_dir = Path(__file__).resolve().parent.parent.parent.parent / "data"
    xlsx_file = data_dir / "customers.xlsx"
    
    if not xlsx_file.exists():
        return []
    
    try:
        if load_workbook is not None:
            wb = load_workbook(xlsx_file)
            ws = wb.active
            customers = set()
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:  # Skip header
                    continue
                if row and row[0]:
                    customers.add(str(row[0]).strip())
            return sorted(list(customers))
        else:
            print("Warning: openpyxl not installed")
            return []
    except Exception as e:
        print(f"Error loading customers from XLSX: {e}")
        return []


def get_customers() -> List[str]:
    """Get customers from available source"""
    customers = load_customers_from_ods()
    if not customers:
        customers = load_customers_from_xlsx()
    return customers if customers else []
